"""
Build the Entravision Digital Media Proposal Template (.xlsx) from the AdFlo catalog.

Generates 6 worksheets:
  - Proposal A             (Net only)
  - Proposal A (Gross)     (Net + Gross, used when has_agency=True)
  - Proposal A (wsections) (sectionable variant for complex proposals)
  - Avails-Only            (potential reach without a billable plan)
  - DOOH Summary           (DOOH-specific rollup)
  - DOOH Screenlist        (DOOH-specific screen-level detail)
  - Process FAQs           (resource links)
"""

import argparse
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from app.catalog import CATALOG, Product, families, by_family
from app.services.notion_parser import compose_target_fallback

# Try to import Image support for logo embedding (optional)
try:
    from openpyxl.drawing.image import Image as XLImage
    _HAS_PIL = True
except Exception:
    _HAS_PIL = False

# Path to the white logo SVG/PNG (optional; used in Excel header if present)
_STATIC_DIR = Path(__file__).resolve().parent / "static"
_LOGO_WHITE_PNG = _STATIC_DIR / "entravision-logo-white.png"
_LOGO_WHITE_SVG = _STATIC_DIR / "entravision-logo-white.svg"


# -----------------------------------------------------------------------------
# Styling
# -----------------------------------------------------------------------------

ENTRAVISION_PURPLE = "FF910E95"
LIGHT_GREY = "FFE7E7E7"
DARK_GREY = "FF808080"
WHITE = "FFFFFFFF"
LIGHT_PURPLE_BG = "FFEFE0EF"
LIGHT_GREEN_BG = "FFE8F2DD"

TITLE_FONT = Font(name="Calibri", size=17, bold=False, color=WHITE)
TITLE_FILL = PatternFill("solid", start_color=ENTRAVISION_PURPLE)
HEADER_PURPLE_FILL = PatternFill("solid", start_color=ENTRAVISION_PURPLE)

H_BRAND = Font(name="Calibri", size=11, bold=True, color=ENTRAVISION_PURPLE)
H_HEADER = Font(name="Tahoma", size=10, bold=True, color=WHITE)
H_HEADER_FILL = PatternFill("solid", start_color=DARK_GREY)

BODY_FONT = Font(name="Arial", size=10, bold=False)
BODY_BOLD = Font(name="Arial", size=10, bold=True)
NOTE_FONT = Font(name="Arial", size=9, color="FF555555", italic=True)
SECTION_FONT = Font(name="Arial", size=12, bold=True, color=WHITE)
SECTION_FILL = PatternFill("solid", start_color="FF7A0F7F")
SUBTOTAL_FONT = Font(name="Arial", size=11, bold=True)
SUBTOTAL_FILL = PatternFill("solid", start_color=LIGHT_PURPLE_BG)
TOTAL_FONT = Font(name="Arial", size=12, bold=True, color="FF000000")
TOTAL_FILL = PatternFill("solid", start_color=LIGHT_GREEN_BG)

THIN = Side(border_style="thin", color="FFBBBBBB")
MED = Side(border_style="medium", color="FF000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THIN, right=THIN, top=MED, bottom=MED)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


# -----------------------------------------------------------------------------
# Column layouts
# Net: A B C D E F G H I J K L M N O P Q R S T U V
# Gross: A B C D E F G H I J K L M N O P Q R S T U V W X Y
# -----------------------------------------------------------------------------


def _apply_col_widths(ws: Worksheet, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _set_header(ws: Worksheet, row: int, headers: list) -> None:
    for col, label in headers:
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = H_HEADER
        c.fill = H_HEADER_FILL
        c.alignment = CENTER
        c.border = HEADER_BORDER
    ws.row_dimensions[row].height = 32


def _write_meta_block(ws: Worksheet, title: str, include_billing: bool = True,
                      include_campaign_meta: bool = True) -> None:
    """
    Top brand block. Rows 1-15 (compacted — content stops well before the
    row-17 header so the block doesn't crowd the visible screen).
    C2:D2 — Entravision purple with white logo.
    E2:L2 — title on purple.

    include_billing: set False for sheets (e.g. Avails-Only) where a customer
    billing block doesn't apply.
    include_campaign_meta: set False for sheets whose product rows start at
    row 11 (e.g. Avails-Only) — the C11:C14 "Media Proposal / Target / Geo"
    lines would otherwise land right on top of the first product rows.
    """
    # C2 + D2: purple background (logo area)
    for col in ("C", "D"):
        cell = ws[f"{col}2"]
        cell.fill = HEADER_PURPLE_FILL

    _insert_logo(ws)

    # E2:L2 title
    ws["E2"] = title
    ws["E2"].font = Font(name="Calibri", size=17, bold=False, color=WHITE)
    ws["E2"].fill = TITLE_FILL
    ws["E2"].alignment = RIGHT
    ws.merge_cells("E2:L2")
    ws.row_dimensions[2].height = 40  # taller header row — gives the logo room to fill it

    # Left column: Entravision contact info (tightened row heights — rows 4-10)
    ws["C4"] = "ENTRAVISION COMMUNICATIONS CORPORATION"
    ws["C4"].font = H_BRAND
    ws["C5"] = "Entravision Contact Information"
    ws["C5"].font = BODY_BOLD
    ws["C6"] = "1 Estrella Way"
    ws["C7"] = "Burbank, CA 91504"
    for r in range(4, 11):
        ws.row_dimensions[r].height = 14

    # Right column: Customer billing (skipped where not relevant, e.g. Avails-Only)
    if include_billing:
        ws["F4"] = "CUSTOMER BILLING INFORMATION:"
        ws["F4"].font = H_BRAND
        ws["F5"] = "Attn: "
        ws["F6"] = "Billing Name: "
        ws["F7"] = "Contact e-mail: "
        ws["F8"] = "Address: "

    if not include_campaign_meta:
        return

    # Campaign meta
    ws["C11"] = "Media Proposal: "
    ws["C11"].font = BODY_BOLD
    ws["C12"] = "Order Description: "
    ws["C12"].font = BODY_BOLD
    ws["C13"] = "Geo: "
    # Combined onto one line to save vertical space (was two separate rows)
    ws["C14"] = "All rates are NET. Minimum 3 month Commitment."
    ws["C14"].font = BODY_BOLD
    ws.row_dimensions[15].height = 6  # spacer row before the avails banner/header


def _insert_logo(ws: Worksheet) -> None:
    """
    Embed the white Entravision logo at C2 if Pillow + the asset are available;
    otherwise fall back to a plain text wordmark so the sheet never breaks.
    """
    if _HAS_PIL and _LOGO_WHITE_PNG.exists():
        try:
            img = XLImage(str(_LOGO_WHITE_PNG))
            # Source is ~1772x517 (3.43:1). Row 2 is 40pt (~53px) tall — size
            # the logo to fill nearly the full row height (C2:D2 is far wider
            # than the logo needs, so height is the binding dimension here).
            img.height = 48
            img.width = 48 * 3.43
            img.anchor = "C2"
            ws.add_image(img)
            return
        except Exception:
            pass  # fall through to text wordmark

    ws["C2"] = "ENTRAVISION"
    ws["C2"].font = Font(name="Calibri", size=15, bold=True, color=WHITE)
    ws["C2"].alignment = CENTER
    ws.merge_cells("C2:D2")


def _format_money_cell(cell, blue_input: bool = False) -> None:
    cell.number_format = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    cell.font = Font(name="Arial", size=10, bold=False, color="FF0000FF" if blue_input else "FF000000")
    cell.alignment = CENTER


def _format_imps_cell(cell) -> None:
    cell.number_format = '#,##0;[Red](#,##0);"-"'
    cell.font = BODY_FONT
    cell.alignment = CENTER


def _set_array_formula(ws: Worksheet, cell_ref: str, formula_text: str) -> None:
    """
    Write a formula as an explicit legacy array formula instead of a plain
    string. A plain `ws[ref] = "=SUMPRODUCT(IFERROR(range*1,0))"` still gets
    an "@" implicit-intersection operator auto-inserted by Excel in front of
    the range wherever it sits inside IFERROR() — even nested inside
    SUMPRODUCT, which doesn't protect its own nested calls from that
    heuristic — because the file has no array-formula marker telling Excel
    the formula is exempt. Explicitly marking it as an array formula (what
    Ctrl+Shift+Enter used to do) is what actually suppresses that rewrite.
    """
    ws[cell_ref] = ArrayFormula(ref=cell_ref, text=formula_text)


def compute_sov_pct(product: Product, monthly_budget: float, avail: dict) -> Optional[float]:
    """
    % of the planner-entered avails ceiling that the curated monthly budget
    would consume — the "Share of Voice" figure: how much of the available
    inventory this campaign is claiming at that product's rate. Compared
    against max_spend when given; falls back to deriving an equivalent spend
    from max_imps via the product's own rate model when spend isn't entered.
    """
    if not monthly_budget:
        return None

    max_spend = avail.get("max_spend")
    if max_spend:
        return monthly_budget / max_spend * 100

    max_imps = avail.get("max_imps")
    if not max_imps:
        return None

    rate = product.base_rate
    if product.buying_model == "CPM" and rate:
        implied_spend = max_imps * rate / 1000
    elif product.buying_model == "CPP" and rate:
        implied_spend = max_imps * rate
    elif product.estimated_cpm_for_imps:
        implied_spend = max_imps * product.estimated_cpm_for_imps / 1000
    else:
        return None
    if not implied_spend:
        return None
    return monthly_budget / implied_spend * 100


SOV_RED = "FFF8D7DA"
SOV_ORANGE = "FFFDE3CF"
SOV_YELLOW = "FFFFF3CD"
SOV_GREEN = LIGHT_GREEN_BG


def _sov_fill(pct: float) -> PatternFill:
    """Traffic-light fill matching the Step 06 avails UI: <65 green, 65-79
    yellow, 80-89 orange, >=90 red."""
    if pct >= 90:
        color = SOV_RED
    elif pct >= 80:
        color = SOV_ORANGE
    elif pct >= 65:
        color = SOV_YELLOW
    else:
        color = SOV_GREEN
    return PatternFill("solid", start_color=color)


def _apply_sov_conditional_formatting(ws: Worksheet, col: str, first_row: int, last_row: int) -> None:
    """
    Traffic-light conditional formatting for the SOV% column, mirroring
    `_sov_fill`'s thresholds. The fill `write_avails_cells` sets on each cell
    is just the as-generated snapshot; once SOV becomes a live formula (see
    below), the planner can change the budget or avails numbers right in
    Excel afterward, and only real conditional formatting — not a static
    PatternFill baked in at generation time — keeps the color honest as the
    formula's result changes. Rules are evaluated in the order added (first
    = highest priority); stopIfTrue keeps a looser rule from also firing
    once a tighter one already matched.
    """
    if first_row > last_row:
        return
    rng = f"{col}{first_row}:{col}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.9"],
                                                   fill=PatternFill("solid", start_color=SOV_RED), stopIfTrue=True))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
                                                   fill=PatternFill("solid", start_color=SOV_ORANGE), stopIfTrue=True))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.65"],
                                                   fill=PatternFill("solid", start_color=SOV_YELLOW), stopIfTrue=True))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.65"],
                                                   fill=PatternFill("solid", start_color=SOV_GREEN), stopIfTrue=True))


def _spend_formula_from_imps(product: Product, imps_cell: str, rate_cell: str,
                             est_cpm_cell: Optional[str]) -> Optional[str]:
    """Live formula for Max. Recommended Monthly Spend, derived from the Max.
    Recommended Monthly Imps cell — so the spend recalculates automatically
    if the planner edits the imps number (or the rate, when the rate itself
    lives in a cell) after the file is generated."""
    if product.buying_model == "CPM" and product.base_rate is not None:
        return f'=IFERROR({imps_cell}*{rate_cell}/1000,"")'
    if product.buying_model == "CPP" and product.base_rate is not None:
        return f'=IFERROR({imps_cell}*{rate_cell},"")'
    if (product.buying_model == "Fixed" or product.estimated_impressions) and product.estimated_cpm_for_imps:
        cpm_ref = est_cpm_cell or product.estimated_cpm_for_imps
        return f'=IFERROR("Est. $"&TEXT({imps_cell}*{cpm_ref}/1000,"#,##0"),"")'
    return None


def _imps_formula_from_spend(product: Product, spend_cell: str, rate_cell: str,
                             est_cpm_cell: Optional[str]) -> Optional[str]:
    """Live formula for Max. Recommended Monthly Imps, derived from the Max.
    Recommended Monthly Spend cell — the mirror image of
    `_spend_formula_from_imps`."""
    if product.buying_model == "CPM" and product.base_rate is not None:
        return f'=IFERROR({spend_cell}/{rate_cell}*1000,0)'
    if product.buying_model == "CPP" and product.base_rate is not None:
        return f'=IFERROR({spend_cell}/{rate_cell},0)'
    if (product.buying_model == "Fixed" or product.estimated_impressions) and product.estimated_cpm_for_imps:
        cpm_ref = est_cpm_cell or product.estimated_cpm_for_imps
        return f'=IFERROR("Est. "&TEXT({spend_cell}*1000/{cpm_ref},"#,##0"),"")'
    return None


def write_avails_cells(ws: Worksheet, row: int, avail: dict, product: Optional[Product] = None,
                       gross: bool = False, cols: Optional[tuple] = None, sov_pct: Optional[float] = None,
                       sov_col: Optional[str] = None, budget_col: Optional[str] = None,
                       rate_col: str = "K", est_cpm_col: Optional[str] = None) -> None:
    """
    Write planner-entered avails (from the app's Step 06) into the
    Max. Recommended Monthly Imps / Spend / Est. Monthly Uniques columns.
    Net sheets: N/O/P. Gross sheets: P/Q/R. Pass `cols=("I","J","K")` for the
    Avails-Only sheet, which uses a different column layout.

    `avail` shape: {max_imps, max_spend, est_uniques, basis?,
                    max_imps_estimated?, max_spend_estimated?}. `basis` is
    "imps" or "spend" — whichever field the planner directly typed most
    recently in Step 06 (the other one is always mechanically derived from
    it there). When `product` + `basis` are given, the derived side is
    written as a live formula referencing the typed cell (and the rate
    cell/`rate_col`, or `est_cpm_col` when the product exposes one) instead
    of a static number, so it keeps recalculating if the planner edits
    either number after the file is generated — that's also what lets the
    SOV formula below stay accurate for a typed-or-derived NUMERIC spend.
    Legacy avails saved before `basis` existed (or a product with no
    imps<->spend conversion) fall back to the old static-number behavior.

    A Fixed-rate/estimated-CPM product's derived value is still written as
    "Est. …" text (now via a formula that evaluates to text, same as
    before) — informational, and intentionally excluded from SUM totals
    like every other "Est." cell in this template, whether static or live.

    sov_pct / sov_col: the "Share of Voice" figure (see compute_sov_pct) and
    the column to write it into. When `budget_col` is given and the Max
    Spend cell holds/will hold a genuine number (not "Est." text), SOV is
    written as a live `=budget/spend` formula instead — it recalculates if
    the planner edits either cell in Excel. Falls back to the pre-computed
    static `sov_pct` otherwise (the Fixed/estimated-text case, or no
    `budget_col` given, e.g. legacy callers).
    """
    if cols:
        imps_col, spend_col, uniques_col = cols
    else:
        imps_col, spend_col, uniques_col = ("P", "Q", "R") if gross else ("N", "O", "P")

    imps_cell_ref = f"{imps_col}{row}"
    spend_cell_ref = f"{spend_col}{row}"
    rate_cell_ref = f"{rate_col}{row}"
    est_cpm_cell_ref = f"{est_cpm_col}{row}" if est_cpm_col else None

    basis = avail.get("basis")
    max_imps = avail.get("max_imps")
    max_spend = avail.get("max_spend")

    imps_formula = spend_formula = None
    if product is not None:
        if basis == "imps" and max_imps is not None:
            spend_formula = _spend_formula_from_imps(product, imps_cell_ref, rate_cell_ref, est_cpm_cell_ref)
        elif basis == "spend" and max_spend is not None:
            imps_formula = _imps_formula_from_spend(product, spend_cell_ref, rate_cell_ref, est_cpm_cell_ref)

    spend_is_text = False  # tracked so the SOV formula below knows not to divide by it

    if max_imps is not None or imps_formula:
        cell = ws[imps_cell_ref]
        if imps_formula:
            cell.value = imps_formula
            cell.alignment = CENTER
            cell.font = BODY_FONT
        elif avail.get("max_imps_estimated"):
            cell.value = f"Est. {max_imps:,.0f}"
            cell.alignment = CENTER
            cell.font = BODY_FONT
        else:
            cell.value = max_imps
            _format_imps_cell(cell)
        _merge_thin_border(cell)

    if max_spend is not None or spend_formula:
        cell = ws[spend_cell_ref]
        if spend_formula:
            cell.value = spend_formula
            cell.alignment = CENTER
            cell.font = BODY_FONT
            spend_is_text = "Est. $" in spend_formula
        elif avail.get("max_spend_estimated"):
            cell.value = f"Est. ${max_spend:,.0f}"
            cell.alignment = CENTER
            cell.font = BODY_FONT
            spend_is_text = True
        else:
            cell.value = max_spend
            _format_money_cell(cell)
        _merge_thin_border(cell)

    est_uniques = avail.get("est_uniques")
    if est_uniques is not None:
        cell = ws[f"{uniques_col}{row}"]
        cell.value = est_uniques
        _format_imps_cell(cell)
        _merge_thin_border(cell)

    if sov_col:
        cell = ws[f"{sov_col}{row}"]
        if budget_col and (max_spend is not None or spend_formula) and not spend_is_text:
            cell.value = f'=IFERROR({budget_col}{row}/{spend_cell_ref},"")'
        elif sov_pct is not None:
            cell.value = round(sov_pct, 1) / 100
        else:
            cell.value = None
        if cell.value is not None:
            cell.number_format = "0.0%"
            cell.font = BODY_FONT
            cell.alignment = CENTER
            if sov_pct is not None:
                cell.fill = _sov_fill(sov_pct)
            _merge_thin_border(cell)


def _estimate_row_height(text: str, col_width: float, *, min_height: float = 45, extra_lines: int = 0) -> float:
    """
    Approximate a wrap-friendly row height for a text-heavy cell (e.g. the AI
    blurb in column E, or a two-line Primary/Secondary target) so long
    descriptions aren't clipped or overly cramped.
    Rough heuristic: ~1.8 characters per column-width-unit per line, ~14pt per line.

    Sums the wrapped-line estimate PER embedded "\\n" segment rather than
    across the whole string — a hard line break always forces at least one
    extra visual line regardless of the surrounding text's total character
    count, which a single whole-string char count/chars_per_line division
    would under-count whenever the combined text is short.
    """
    if not text:
        return min_height
    chars_per_line = max(int(col_width * 1.8), 10)
    lines = sum(-(-max(len(segment), 1) // chars_per_line) for segment in text.split("\n"))
    lines += extra_lines
    height = lines * 14 + 16  # + padding
    return max(min_height, height)


def _merge_thin_border(cell) -> None:
    """
    Apply THIN_BORDER to `cell` without ever downgrading a side that's
    already MEDIUM (e.g. the outer "ticket" box _apply_outer_border draws
    around the line-item block). Avails are written to their cells AFTER
    that box is drawn, and a plain `cell.border = THIN_BORDER` there would
    silently erase the medium edge wherever an avails value lands on the
    perimeter — this preserves it instead.
    """
    existing = cell.border

    def pick(side):
        return side if (side is not None and side.style == "medium") else THIN

    cell.border = Border(
        left=pick(existing.left), right=pick(existing.right),
        top=pick(existing.top), bottom=pick(existing.bottom),
    )


def _box_range(ws: Worksheet, min_row: int, max_row: int, min_col: str, max_col: str) -> None:
    """
    Give a range a clean grid — thin borders between every cell — plus a
    thick "ticket" outline around its perimeter. Used for the totals rows
    and the add-ons table so they read as the same kind of grouped block as
    the product line-items table, on an otherwise plain white sheet.
    """
    from openpyxl.utils import column_index_from_string
    min_c = column_index_from_string(min_col)
    max_c = column_index_from_string(max_col)
    for r in range(min_row, max_row + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if cell.border.left.style is None:
                cell.border = THIN_BORDER
    _apply_outer_border(ws, min_row, max_row, min_col, max_col)


def _fill_box_range(ws: Worksheet, min_row: int, max_row: int, min_col: str, max_col: str, fill: PatternFill) -> None:
    """
    Apply `fill` to every cell in a range — used so the totals rows' green
    highlight covers the full outlined row instead of just the handful of
    cells that happen to hold a value, which read as a patchy highlight.
    """
    from openpyxl.utils import column_index_from_string
    min_c = column_index_from_string(min_col)
    max_c = column_index_from_string(max_col)
    for r in range(min_row, max_row + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(row=r, column=c).fill = fill


def _apply_outer_border(ws: Worksheet, min_row: int, max_row: int, min_col: str, max_col: str) -> None:
    """
    Draw a medium-weight box around a range of cells (e.g. the product line-item
    block) so it reads as a single grouped section — a "ticket" outline —
    on top of each cell's own thin grid border.
    """
    from openpyxl.utils import column_index_from_string
    min_c = column_index_from_string(min_col)
    max_c = column_index_from_string(max_col)

    for r in range(min_row, max_row + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            left = MED if c == min_c else existing.left
            right = MED if c == max_c else existing.right
            top = MED if r == min_row else existing.top
            bottom = MED if r == max_row else existing.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


_PRIMARY_LABEL_FONT = InlineFont(b=True, color=ENTRAVISION_PURPLE)
_SECONDARY_LABEL_FONT = InlineFont(b=True, color=DARK_GREY)


def build_target_cell_value(target_override: Optional[str], target_secondary: Optional[str], request):
    """
    Build column D's TARGET cell value. When the planner has added a
    secondary audience (used for scale — e.g. broadening a narrow primary
    intent segment to pull more avails), the cell becomes a two-line rich-
    text value:
        Primary: <the primary target>
        Secondary: <the secondary target>
    with the "Primary:"/"Secondary:" labels bolded and colored for
    clarity — this is the most Excel can do for "highlighting" specific
    words inside one cell; a background fill can only apply to the whole
    cell, not individual runs of text within it.

    With no secondary audience, returns exactly what this column has always
    returned: the per-line override, or the campaign-level Demo |
    Behavioral | Contextual fallback — unchanged from before this feature.
    """
    primary = target_override if target_override else compose_target_fallback(request)
    if not target_secondary:
        return primary
    return CellRichText(
        TextBlock(_PRIMARY_LABEL_FONT, "Primary: "), primary, "\n",
        TextBlock(_SECONDARY_LABEL_FONT, "Secondary: "), target_secondary,
    )


# -----------------------------------------------------------------------------
# Sheet builders
# -----------------------------------------------------------------------------


def build_proposal_a(wb: Workbook, products: list, with_sections: bool = False,
                     start_date: str = "", end_date: str = "", total_months: int = 3,
                     sheet_name: Optional[str] = None) -> Worksheet:
    """Build the Net-only proposal sheet. Returns the worksheet.

    sheet_name: override the default "Proposal A" / "Proposal A (wsections)"
    name — used for tiered-budget proposals, where each option gets its own
    lettered tab ("Proposal A", "Proposal B", ...).
    """
    if sheet_name is None:
        sheet_name = "Proposal A (wsections)" if with_sections else "Proposal A"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables

    widths = {
        "A": 1.5, "B": 2.0, "C": 38, "D": 32, "E": 50, "F": 12, "G": 12,
        "H": 20, "I": 16, "J": 11, "K": 12, "L": 14, "M": 2.5,
        "N": 21, "O": 21, "P": 21, "Q": 16, "R": 4, "S": 4, "T": 60, "U": 5, "V": 22,
    }
    _apply_col_widths(ws, widths)

    _write_meta_block(ws, "Digital Plan / Avails")

    # Months input cell (I10) — blue = planner can adjust
    ws["H10"] = "Months:"
    ws["H10"].alignment = RIGHT
    ws["H10"].font = BODY_BOLD
    ws["I10"] = total_months
    ws["I10"].font = Font(name="Arial", size=11, bold=True, color="FF0000FF")
    ws["I10"].alignment = CENTER

    # Avails section banner (row 16)
    ws["N16"] = "Monthly Forecast — (4-week) Avails"
    ws["N16"].font = BODY_BOLD
    ws["N16"].alignment = CENTER
    ws.merge_cells("N16:Q16")

    # Header row (row 17)
    _set_header(ws, 17, [
        ("C", "LINE NAME"),
        ("D", "TARGET"),
        ("E", "DETAILS"),
        ("F", "START DATE"),
        ("G", "END DATE"),
        ("H", "SIZES"),
        ("I", "IMPRESSIONS"),
        ("J", "RATE TYPE"),
        ("K", "NET RATE"),
        ("L", "NET BUDGET"),
        ("N", "Max. Recommended\nMonthly Imps"),
        ("O", "Max. Recommended\nMonthly Spend"),
        ("P", "Est. Monthly Uniques"),
        ("Q", "Est. % of\nAvails Used (SOV)"),
        ("T", "Planner Notes — internal guidance"),
        ("V", "AdOps (Internal Use)"),
    ])
    ws.row_dimensions[17].height = 40  # extra room for 2-line wrapped headers

    # Product rows
    row = 19
    first_data_row = row
    section_idx = 0
    last_family = None
    for p in products:
        if with_sections and p.family != last_family:
            section_idx += 1
            ws[f"C{row}"] = f"SECTION {section_idx}: {p.family.upper()}"
            ws[f"C{row}"].font = SECTION_FONT
            ws[f"C{row}"].fill = SECTION_FILL
            ws[f"C{row}"].alignment = LEFT
            ws.merge_cells(f"C{row}:L{row}")
            ws.row_dimensions[row].height = 22
            row += 1
            last_family = p.family

        _write_product_row(ws, row, p, gross=False,
                           start_date=start_date, end_date=end_date)
        row += 1

    # "Ticket" outline around the whole line-item block for visual grouping
    _apply_outer_border(ws, first_data_row, row - 1, "C", "Q")

    # Totals
    last_data_row = row - 1
    total_row = row + 1
    ws[f"C{total_row}"] = "TOTAL DIGITAL MONTHLY"
    ws[f"C{total_row}"].font = TOTAL_FONT
    ws[f"C{total_row}"].fill = TOTAL_FILL
    ws[f"L{total_row}"] = f"=ROUNDDOWN(SUM(L19:L{last_data_row}),0)"
    _format_money_cell(ws[f"L{total_row}"])
    ws[f"L{total_row}"].font = TOTAL_FONT
    ws[f"L{total_row}"].fill = TOTAL_FILL
    _set_array_formula(ws, f"I{total_row}", f"=SUMPRODUCT(IFERROR(I19:I{last_data_row}*1,0))")
    _format_imps_cell(ws[f"I{total_row}"])
    ws[f"I{total_row}"].font = TOTAL_FONT
    ws[f"I{total_row}"].fill = TOTAL_FILL
    ws[f"J{total_row}"] = '="eCPM:"'
    ws[f"K{total_row}"] = f"=IFERROR(L{total_row}/I{total_row}*1000,\"\")"
    _format_money_cell(ws[f"K{total_row}"])
    _box_range(ws, total_row, total_row, "C", "L")
    _fill_box_range(ws, total_row, total_row, "C", "L", TOTAL_FILL)

    _write_addons_grand_total_footer(ws, total_row, gross=False, box_max_col="L")

    ws.freeze_panes = "C18"
    return ws


def _write_addons_grand_total_footer(ws: Worksheet, total_row: int, *, gross: bool,
                                     box_max_col: str, months_cell: str = "I10") -> int:
    """
    Shared by build_proposal_a and build_proposal_a_gross: the ADD-ONS /
    ONE-TIME FEES block, the campaign-length grand total row, the footer
    copy, and the signature block. The Gross sheet previously stopped right
    after its "TOTAL DIGITAL MONTHLY" row and never got any of this — this
    fixes that by having both sheets build it from one place, so they can't
    drift out of sync again.

    Gross sheets additionally carry a GROSS amount per add-on (column N,
    derived from the NET amount via the $I$14 agency-fee cell) and a GROSS
    grand total alongside the NET one.

    Returns the signature dotted-line row.
    """
    addons_start = total_row + 2
    ws[f"C{addons_start}"] = "ADD-ONS / ONE-TIME FEES"
    ws[f"C{addons_start}"].font = BODY_BOLD
    ws[f"C{addons_start}"].fill = SUBTOTAL_FILL
    ws.merge_cells(f"C{addons_start}:{box_max_col}{addons_start}")
    addons = [
        # Search - SEM Setup fee removed — waived, no longer active (was
        # "Search - SEM Setup — One-time fee (new customers only)", $50 Fixed).
        ("Call Tracking Service", "Tracking phone number with call recording capabilities.", 0.0, "Added Value"),
        ("Online Attribution Measurement", "Attribution and reporting layer.", 0.0, "Added Value"),
        ("Email Database Match — Hashed File Onboarding", "Upload your email database to match opted-in users in our database for precise targeting.", 150.0, "Fixed"),
        ("Web Services — Landing Page", "Optimized landing page with hosting and dedicated URL.", 100.0, "Fixed"),
    ]
    for i, (svc, desc, amt, rtype) in enumerate(addons, start=1):
        r = addons_start + i
        ws[f"C{r}"] = svc
        ws[f"C{r}"].font = BODY_FONT
        ws[f"C{r}"].alignment = LEFT
        ws[f"E{r}"] = desc
        ws[f"E{r}"].alignment = LEFT
        ws[f"E{r}"].font = BODY_FONT
        ws[f"J{r}"] = rtype
        ws[f"J{r}"].alignment = CENTER
        ws[f"L{r}"] = amt
        _format_money_cell(ws[f"L{r}"], blue_input=True)
        if gross:
            ws[f"N{r}"] = f"=IFERROR(L{r}/(1-$I$14),0)"
            _format_money_cell(ws[f"N{r}"])

    addons_last = addons_start + len(addons)
    _box_range(ws, addons_start, addons_last, "C", box_max_col)

    # Grand total — dynamic label uses the months cell
    grand_row = addons_last + 2
    ws[f"C{grand_row}"] = f'="TOTAL DIGITAL — "&{months_cell}&"-MONTH CAMPAIGN"'
    ws[f"C{grand_row}"].font = TOTAL_FONT
    ws[f"C{grand_row}"].fill = TOTAL_FILL
    ws[f"L{grand_row}"] = (
        f"=ROUNDDOWN(L{total_row}*{months_cell}"
        f" + SUMIF(L{addons_start+1}:L{addons_last},\">0\")"
        f",0)"
    )
    _format_money_cell(ws[f"L{grand_row}"])
    ws[f"L{grand_row}"].font = TOTAL_FONT
    ws[f"L{grand_row}"].fill = TOTAL_FILL
    if gross:
        ws[f"N{grand_row}"] = (
            f"=ROUNDDOWN(N{total_row}*{months_cell}"
            f" + SUMIF(N{addons_start+1}:N{addons_last},\">0\")"
            f",0)"
        )
        _format_money_cell(ws[f"N{grand_row}"])
        ws[f"N{grand_row}"].font = TOTAL_FONT
        ws[f"N{grand_row}"].fill = TOTAL_FILL
    _box_range(ws, grand_row, grand_row, "C", box_max_col)
    _fill_box_range(ws, grand_row, grand_row, "C", box_max_col, TOTAL_FILL)

    # Footer
    foot = grand_row + 2
    ws[f"C{foot}"] = (
        "To maximize the efficiency of your digital media investment, we strongly recommend "
        "installing a Google Tag Manager container on your website."
    )
    ws[f"C{foot}"].font = NOTE_FONT
    ws[f"C{foot}"].alignment = LEFT
    ws.merge_cells(f"C{foot}:{box_max_col}{foot}")
    ws[f"C{foot+1}"] = "Payment Terms: Cash in Advance or Net 30 on approved credit"
    ws[f"E{foot+1}"] = "IAB standard terms & conditions apply to all digital media purchases."
    ws[f"C{foot+2}"] = (
        "This proposal will be valid for a period of 1 month after being presented. Please notify "
        "your Account Executive if you require the presented media to remain booked after that time."
    )
    ws.merge_cells(f"C{foot+2}:{box_max_col}{foot+2}")
    ws[f"C{foot+3}"] = "Client accepts Entravision's Terms of Sales (https://entravision.com/termsofsales/)"

    # Signature block
    sig = foot + 5
    ws[f"C{sig}"] = "Customer Signature"
    ws[f"E{sig}"] = "Name"
    ws[f"F{sig}"] = "Title"
    ws[f"C{sig+1}"] = "……………………………………………………..…"
    ws[f"E{sig+1}"] = "……………………………………………………………."
    ws[f"F{sig+1}"] = "……………"
    # Bottom-align the dotted line and give its row real height, so there's
    # actual blank space above the line to physically sign when printed —
    # not just the line itself squeezed into a default-height row.
    for col in ("C", "E", "F"):
        ws[f"{col}{sig+1}"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[sig + 1].height = 40

    return sig + 1


def _write_product_row(ws: Worksheet, row: int, p: Product, gross: bool = False,
                       start_date: str = "", end_date: str = "") -> None:
    """Write a single product line."""
    # SERVICES (C)
    services_label = f"{p.name}\n{p.short_label}"
    ws[f"C{row}"] = services_label
    ws[f"C{row}"].font = BODY_BOLD
    ws[f"C{row}"].alignment = CENTER
    ws[f"C{row}"].border = THIN_BORDER

    # TARGET (D)
    ws[f"D{row}"] = "TBD"
    ws[f"D{row}"].alignment = CENTER
    ws[f"D{row}"].border = THIN_BORDER

    # DETAILS (E)
    ws[f"E{row}"] = p.proposal_description
    ws[f"E{row}"].font = BODY_FONT
    ws[f"E{row}"].alignment = LEFT
    ws[f"E{row}"].border = THIN_BORDER

    # START / END DATE (F, G) — use actual flight dates if available
    ws[f"F{row}"] = start_date or "TBD"
    ws[f"F{row}"].alignment = CENTER
    ws[f"F{row}"].border = THIN_BORDER

    ws[f"G{row}"] = end_date or "TBD"
    ws[f"G{row}"].alignment = CENTER
    ws[f"G{row}"].border = THIN_BORDER

    # SIZES (H)
    ws[f"H{row}"] = p.sizes
    ws[f"H{row}"].alignment = CENTER
    ws[f"H{row}"].border = THIN_BORDER

    # IMPRESSIONS (I) — formula by buying model
    # CPM:   =L/K*1000
    # CPP:   =L/K
    # Fixed with estimated CPM: ="Est. "&TEXT(L*1000/est_cpm,"#,##0")
    # else:  NA
    if p.buying_model == "CPM" and p.base_rate is not None:
        ws[f"I{row}"] = f"=IFERROR(L{row}/K{row}*1000,0)"
        _format_imps_cell(ws[f"I{row}"])
    elif p.buying_model == "CPP" and p.base_rate is not None:
        ws[f"I{row}"] = f"=IFERROR(L{row}/K{row},0)"
        _format_imps_cell(ws[f"I{row}"])
    elif (p.buying_model == "Fixed" or p.estimated_impressions) and p.estimated_cpm_for_imps:
        ws[f"I{row}"] = (
            f'="Est. "&TEXT(L{row}*1000/{p.estimated_cpm_for_imps},"#,##0")'
        )
        ws[f"I{row}"].alignment = CENTER
    else:
        ws[f"I{row}"] = "NA"
        ws[f"I{row}"].alignment = CENTER
    ws[f"I{row}"].border = THIN_BORDER

    # RATE TYPE (J)
    rate_type_display = "Fixed" if (p.estimated_impressions or p.base_rate is None) else p.buying_model
    ws[f"J{row}"] = rate_type_display
    ws[f"J{row}"].alignment = CENTER
    ws[f"J{row}"].border = THIN_BORDER

    # NET RATE (K) — blue = planner input
    if p.base_rate is not None and not p.estimated_impressions:
        ws[f"K{row}"] = p.base_rate
        _format_money_cell(ws[f"K{row}"], blue_input=True)
    else:
        ws[f"K{row}"] = "NA"
        ws[f"K{row}"].alignment = CENTER
        ws[f"K{row}"].font = BODY_FONT
    ws[f"K{row}"].border = THIN_BORDER

    # NET BUDGET (L) — preload with min spend; planner's primary input
    ws[f"L{row}"] = p.minimum_spend
    _format_money_cell(ws[f"L{row}"], blue_input=True)
    ws[f"L{row}"].border = THIN_BORDER

    if gross:
        if p.base_rate is not None and not p.estimated_impressions:
            ws[f"M{row}"] = f'=IF(K{row}<>"NA",K{row}/(1-$I$14),"NA")'
            _format_money_cell(ws[f"M{row}"])
        else:
            ws[f"M{row}"] = "NA"
            ws[f"M{row}"].alignment = CENTER
        ws[f"M{row}"].border = THIN_BORDER
        ws[f"N{row}"] = f"=L{row}/(1-$I$14)"
        _format_money_cell(ws[f"N{row}"])
        ws[f"N{row}"].border = THIN_BORDER

    # Avails columns — blank for planner input
    avails_cols = ("P", "Q", "R") if gross else ("N", "O", "P")
    for col in avails_cols:
        ws[f"{col}{row}"].border = THIN_BORDER
        ws[f"{col}{row}"].alignment = CENTER

    # Notes column
    notes_col = "W" if gross else "T"
    ws[f"{notes_col}{row}"] = p.notes or ""
    ws[f"{notes_col}{row}"].font = NOTE_FONT
    ws[f"{notes_col}{row}"].alignment = LEFT

    # Row height auto-adjusts to the DETAILS (E) text length, with a touch of
    # breathing room, instead of a one-size-fits-all fixed height.
    ws.row_dimensions[row].height = _estimate_row_height(p.proposal_description, col_width=50)


def build_proposal_a_gross(wb: Workbook, products: list,
                           start_date: str = "", end_date: str = "", total_months: int = 3,
                           sheet_name: Optional[str] = None) -> Worksheet:
    """Build the Gross variant.

    sheet_name: override the default "Proposal A (Gross)" — used for
    tiered-budget proposals ("Proposal B (Gross)", etc.).
    """
    ws = wb.create_sheet(sheet_name or "Proposal A (Gross)")
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables

    widths = {
        "A": 1.5, "B": 2.0, "C": 38, "D": 32, "E": 50, "F": 12, "G": 12,
        "H": 20, "I": 16, "J": 11, "K": 12, "L": 14, "M": 12, "N": 14, "O": 2.5,
        "P": 21, "Q": 21, "R": 21, "S": 16, "T": 4, "U": 4, "V": 4, "W": 60, "X": 5, "Y": 22,
    }
    _apply_col_widths(ws, widths)

    _write_meta_block(ws, "Digital Plan / Avails (Gross)")

    # Months input
    ws["H10"] = "Months:"
    ws["H10"].alignment = RIGHT
    ws["H10"].font = BODY_BOLD
    ws["I10"] = total_months
    ws["I10"].font = Font(name="Arial", size=11, bold=True, color="FF0000FF")
    ws["I10"].alignment = CENTER

    # Agency fee input (I14)
    ws["C14"] = "Agency Fee:"
    ws["C14"].font = BODY_BOLD
    ws["I14"] = 0.15
    ws["I14"].number_format = "0%"
    ws["I14"].font = Font(name="Arial", size=11, bold=True, color="FF0000FF")
    ws["I14"].alignment = CENTER
    ws["I14"].fill = PatternFill("solid", start_color="FFFFFF00")

    ws["P16"] = "Monthly Forecast — (4-week) Avails"
    ws["P16"].font = BODY_BOLD
    ws["P16"].alignment = CENTER
    ws.merge_cells("P16:S16")

    _set_header(ws, 17, [
        ("C", "LINE NAME"),
        ("D", "TARGET"),
        ("E", "DETAILS"),
        ("F", "START DATE"),
        ("G", "END DATE"),
        ("H", "SIZES"),
        ("I", "IMPRESSIONS"),
        ("J", "RATE TYPE"),
        ("K", "NET RATE"),
        ("L", "NET BUDGET"),
        ("M", "GROSS RATE"),
        ("N", "GROSS BUDGET"),
        ("P", "Max. Recommended\nMonthly Imps"),
        ("Q", "Max. Recommended\nMonthly Spend"),
        ("R", "Est. Monthly Uniques"),
        ("S", "Est. % of\nAvails Used (SOV)"),
        ("W", "Planner Notes — internal guidance"),
        ("Y", "AdOps (Internal Use)"),
    ])
    ws.row_dimensions[17].height = 40  # extra room for 2-line wrapped headers

    row = 19
    first_data_row = row
    for p in products:
        _write_product_row(ws, row, p, gross=True,
                           start_date=start_date, end_date=end_date)
        row += 1

    # "Ticket" outline around the whole line-item block for visual grouping
    _apply_outer_border(ws, first_data_row, row - 1, "C", "S")

    last_data_row = row - 1
    total_row = row + 1
    ws[f"C{total_row}"] = "TOTAL DIGITAL MONTHLY"
    ws[f"C{total_row}"].font = TOTAL_FONT
    ws[f"C{total_row}"].fill = TOTAL_FILL
    ws[f"L{total_row}"] = f"=ROUNDDOWN(SUM(L19:L{last_data_row}),0)"
    _format_money_cell(ws[f"L{total_row}"])
    ws[f"L{total_row}"].font = TOTAL_FONT
    ws[f"L{total_row}"].fill = TOTAL_FILL
    ws[f"N{total_row}"] = f"=ROUNDDOWN(SUM(N19:N{last_data_row}),0)"
    _format_money_cell(ws[f"N{total_row}"])
    ws[f"N{total_row}"].font = TOTAL_FONT
    ws[f"N{total_row}"].fill = TOTAL_FILL
    _set_array_formula(ws, f"I{total_row}", f"=SUMPRODUCT(IFERROR(I19:I{last_data_row}*1,0))")
    _format_imps_cell(ws[f"I{total_row}"])
    ws[f"I{total_row}"].font = TOTAL_FONT
    ws[f"I{total_row}"].fill = TOTAL_FILL
    _box_range(ws, total_row, total_row, "C", "N")
    _fill_box_range(ws, total_row, total_row, "C", "N", TOTAL_FILL)

    _write_addons_grand_total_footer(ws, total_row, gross=True, box_max_col="N")

    ws.freeze_panes = "C18"
    return ws


def build_avails_only(wb: Workbook, products: list, *,
                      line_items: Optional[list] = None,
                      request=None,
                      avails_data: Optional[dict] = None,
                      campaign_name: str = "",
                      start_date: str = "", end_date: str = "",
                      sheet_name: Optional[str] = None) -> Worksheet:
    """Avails-only sheet — layout per spec:
    LINE NAME | TARGET | GEO | BUY TYPE | CPM | Est. CPM |
    Max. Recommended Monthly Imps | Max. Recommended Monthly Spend | Est. Monthly Uniques

    `line_items` (planner's LineItem objects, same order as `products`) and
    `request` (the ProposalRequest) drive the per-row TARGET/GEO columns.
    `avails_data` (product_name -> {max_imps, max_spend, est_uniques}) drives
    I/J/K directly when the planner has entered avails in the app; otherwise
    I/J/K fall back to the in-sheet auto-calc formula. `campaign_name` is the
    AI-inferred short campaign name shown in the "Order Description:" line.
    `sheet_name` overrides the default "Avails-Only" — used for tiered-budget
    proposals ("Avails-Only B", etc.).
    """
    ws = wb.create_sheet(sheet_name or "Avails-Only")
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables

    widths = {
        "A": 1.5, "B": 2.0,
        "C": 36,   # LINE NAME
        "D": 30,   # TARGET
        "E": 24,   # GEO
        "F": 12,   # BUY TYPE
        "G": 12,   # CPM
        "H": 12,   # Est. CPM
        "I": 16,   # MONTHLY BUDGET
        "J": 22,   # Max Monthly Imps
        "K": 22,   # Max Monthly Spend
        "L": 22,   # Est. Monthly Uniques
        "M": 16,   # Est. % of Avails Used (SOV)
        "N": 4, "O": 60,  # notes
    }
    _apply_col_widths(ws, widths)

    # No customer billing block, and no C11:C14 campaign-meta lines on this
    # sheet type — its product rows start at row 11, which those lines would
    # otherwise overwrite.
    _write_meta_block(ws, "Avails / Estimates Only", include_billing=False, include_campaign_meta=False)

    client_name = getattr(request, "client_name", "") if request else ""
    campaign_geo = getattr(request, "geo", "") if request else ""

    ws["C5"] = f"Client: {client_name}"
    ws["C5"].font = BODY_BOLD
    ws["C6"] = f"Order Description: {campaign_name}" if campaign_name else "Order Description: "
    ws["C6"].font = BODY_BOLD
    ws["C7"] = "All rates are NET. This is a forecast — not a guarantee of delivery."
    ws["C7"].font = BODY_BOLD

    if start_date or end_date:
        ws["C8"] = f"Flight: {start_date or 'TBD'} – {end_date or 'TBD'}"
        ws["C8"].font = BODY_FONT

    _set_header(ws, 10, [
        ("C", "LINE NAME"),
        ("D", "TARGET"),
        ("E", "GEO"),
        ("F", "BUY TYPE"),
        ("G", "CPM"),
        ("H", "Est. CPM"),
        ("I", "Monthly Budget"),
        ("J", "Max. Recommended\nMonthly Imps"),
        ("K", "Max. Recommended\nMonthly Spend"),
        ("L", "Est. Monthly Uniques"),
        ("M", "Est. % of\nAvails Used (SOV)"),
        ("O", "Planner Notes"),
    ])
    ws.row_dimensions[10].height = 40

    line_items = line_items or []
    row = 12
    first_data_row = row
    for idx, p in enumerate(products):
        li = line_items[idx] if idx < len(line_items) else None

        ws[f"C{row}"] = p.short_label
        ws[f"C{row}"].font = BODY_BOLD
        ws[f"C{row}"].alignment = LEFT
        ws[f"C{row}"].border = THIN_BORDER

        # TARGET (D) — per-line override, else campaign-level fallback; a
        # secondary audience turns this into a two-line Primary/Secondary
        # rich-text cell with bolded labels (see build_target_cell_value).
        target_value = build_target_cell_value(
            li.target_override if li else None, li.target_secondary if li else None, request,
        )
        ws[f"D{row}"] = target_value
        ws[f"D{row}"].alignment = LEFT
        ws[f"D{row}"].border = THIN_BORDER
        # str() of a plain string is a no-op; str() of a CellRichText joins
        # its runs into readable text — either way this is a real char count
        # for row-height sizing below, unlike len() on the rich-text object
        # itself (which would count its internal runs, not characters).
        target_text = str(target_value)

        # GEO (E) — campaign geo (same across all lines; single-campaign flight)
        ws[f"E{row}"] = campaign_geo or "TBD"
        ws[f"E{row}"].alignment = LEFT
        ws[f"E{row}"].border = THIN_BORDER

        ws[f"F{row}"] = "Fixed" if p.estimated_impressions else p.buying_model
        ws[f"F{row}"].alignment = CENTER
        ws[f"F{row}"].border = THIN_BORDER

        ws[f"G{row}"] = p.base_rate if (p.base_rate is not None and not p.estimated_impressions) else "NA"
        if isinstance(ws[f"G{row}"].value, (int, float)):
            _format_money_cell(ws[f"G{row}"], blue_input=True)
        ws[f"G{row}"].alignment = CENTER
        ws[f"G{row}"].border = THIN_BORDER

        # Est. CPM (H) — shown for budget-based/Fixed products that carry one
        ws[f"H{row}"] = p.estimated_cpm_for_imps or "NA"
        if isinstance(ws[f"H{row}"].value, (int, float)):
            _format_money_cell(ws[f"H{row}"], blue_input=True)
        ws[f"H{row}"].alignment = CENTER
        ws[f"H{row}"].border = THIN_BORDER

        # Monthly Budget (I) — the planner's curated spend for this line, shown
        # here (unlike the other columns) so SOV can be a live =budget/max_spend
        # formula on this sheet too, the same as the Net/Gross proposal tabs.
        ws[f"I{row}"] = li.monthly_budget if li else None
        if isinstance(ws[f"I{row}"].value, (int, float)):
            _format_money_cell(ws[f"I{row}"], blue_input=True)
        ws[f"I{row}"].border = THIN_BORDER

        # Keyed by the line item's own id first (falls back to product name)
        # so two lines sharing the same product don't collide on avails.
        avails_by = avails_data or {}
        avail = avails_by.get(li.id) if (li and li.id) else None
        if avail is None:
            avail = avails_by.get(p.name)
        if avail and (avail.get("max_imps") is not None or avail.get("max_spend") is not None):
            # Planner already computed avails in the app (Step 06) — write directly.
            sov_pct = compute_sov_pct(p, li.monthly_budget if li else 0, avail)
            write_avails_cells(ws, row, avail, p, cols=("J", "K", "L"), sov_pct=sov_pct, sov_col="M",
                               budget_col="I", rate_col="G", est_cpm_col="H")
        else:
            # Fallback: leave J open for manual planner input, auto-calc K from it.
            ws[f"J{row}"].border = THIN_BORDER
            ws[f"J{row}"].alignment = CENTER

            if p.buying_model == "CPM" and not p.estimated_impressions:
                ws[f"K{row}"] = f'=IFERROR(J{row}*G{row}/1000,"")'
            elif p.buying_model == "CPP" and not p.estimated_impressions:
                ws[f"K{row}"] = f'=IFERROR(J{row}*G{row},"")'
            elif p.estimated_impressions and p.estimated_cpm_for_imps:
                ws[f"K{row}"] = f'=IFERROR("Est. $"&TEXT(J{row}*H{row}/1000,"#,##0"),"")'
            else:
                ws[f"K{row}"] = f'=IFERROR(J{row}*G{row}/1000,"")'
            _format_money_cell(ws[f"K{row}"])
            ws[f"K{row}"].border = THIN_BORDER

            ws[f"L{row}"].border = THIN_BORDER
            ws[f"L{row}"].alignment = CENTER

            # SOV still works here as a live formula — Monthly Budget (I) over
            # the auto-calc'd Max Spend (K) — even without planner-entered avails.
            ws[f"M{row}"] = f'=IFERROR(I{row}/K{row},"")'
            ws[f"M{row}"].number_format = "0.0%"
            ws[f"M{row}"].font = BODY_FONT
            ws[f"M{row}"].alignment = CENTER
            ws[f"M{row}"].border = THIN_BORDER

        ws[f"O{row}"] = p.notes or ""
        ws[f"O{row}"].font = NOTE_FONT
        ws[f"O{row}"].alignment = LEFT

        ws.row_dimensions[row].height = max(30, _estimate_row_height(target_text, col_width=22, min_height=30))
        row += 1

    # "Ticket" outline around the whole line-item block for visual grouping
    _apply_outer_border(ws, first_data_row, row - 1, "C", "M")
    _apply_sov_conditional_formatting(ws, "M", first_data_row, row - 1)

    # Totals row
    last = row - 1
    total_row = row + 1
    ws[f"C{total_row}"] = "TOTAL DIGITAL — FORECASTED AVAILS / SPEND / UNIQUES"
    ws[f"C{total_row}"].font = TOTAL_FONT
    ws[f"C{total_row}"].fill = TOTAL_FILL
    _set_array_formula(ws, f"I{total_row}", f"=SUMPRODUCT(IFERROR(I{first_data_row}:I{last}*1,0))")
    _set_array_formula(ws, f"J{total_row}", f"=SUMPRODUCT(IFERROR(J{first_data_row}:J{last}*1,0))")
    _set_array_formula(ws, f"K{total_row}", f"=SUMPRODUCT(IFERROR(K{first_data_row}:K{last}*1,0))")
    _set_array_formula(ws, f"L{total_row}", f"=SUMPRODUCT(IFERROR(L{first_data_row}:L{last}*1,0))")
    for col in ("I", "J", "K", "L"):
        ws[f"{col}{total_row}"].fill = TOTAL_FILL
        ws[f"{col}{total_row}"].font = TOTAL_FONT
    _format_money_cell(ws[f"I{total_row}"])
    _format_imps_cell(ws[f"J{total_row}"])
    _format_money_cell(ws[f"K{total_row}"])
    _format_imps_cell(ws[f"L{total_row}"])
    _box_range(ws, total_row, total_row, "C", "M")
    _fill_box_range(ws, total_row, total_row, "C", "M", TOTAL_FILL)

    foot = total_row + 2
    ws[f"C{foot}"] = (
        "*Avails are valid for 1 month after presentation. "
        "This forecast does not constitute a guarantee of delivery."
    )
    ws[f"C{foot}"].font = NOTE_FONT
    ws.merge_cells(f"C{foot}:M{foot}")

    return ws


def build_dooh_summary(wb: Workbook, client_name: str = "") -> Worksheet:
    """DOOH Summary — clean header + placeholder pivot table referencing the Screenlist tab."""
    ws = wb.create_sheet("DOOH Summary")
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables

    widths = {"A": 1.5, "B": 32, "C": 24, "D": 14, "E": 22, "F": 22}
    _apply_col_widths(ws, widths)

    _write_meta_block(ws, f"DOOH Summary{' — ' + client_name if client_name else ''}")

    ws["B4"] = "DOOH Screen Summary"
    ws["B4"].font = BODY_BOLD

    ws["B6"] = "Screens are sourced from the DOOH Screenlist tab."
    ws["B6"].font = NOTE_FONT

    ws["B7"] = "Minimum of 20 screens required."
    ws["B7"].font = BODY_BOLD

    # DOOH CPM assumption (planner input)
    ws["E9"] = "DOOH CPM Assumption:"
    ws["E9"].font = BODY_BOLD
    ws["E9"].alignment = RIGHT
    ws["F9"] = 25.0
    _format_money_cell(ws["F9"], blue_input=True)

    _set_header(ws, 11, [
        ("B", "Venue Type"),
        ("C", "City"),
        ("D", "# Screens"),
        ("E", "Est. Monthly Max. Avails"),
        ("F", "Est. Monthly Max. Spend"),
    ])

    # Two placeholder rows (planner replaces with pivot data from Screenlist)
    for r in (12, 13):
        ws[f"B{r}"] = "—"
        ws[f"C{r}"] = "—"
        ws[f"D{r}"] = 0
        ws[f"E{r}"] = 0
        ws[f"F{r}"] = f"=IFERROR(E{r}*F$9/1000,0)"
        for col in ("B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = THIN_BORDER
        _format_imps_cell(ws[f"D{r}"])
        _format_imps_cell(ws[f"E{r}"])
        _format_money_cell(ws[f"F{r}"])

    # Grand total
    gt_row = 15
    ws[f"B{gt_row}"] = "Grand Total"
    ws[f"B{gt_row}"].font = TOTAL_FONT
    ws[f"B{gt_row}"].fill = TOTAL_FILL
    ws[f"D{gt_row}"] = "=SUM(D12:D14)"
    ws[f"E{gt_row}"] = "=SUM(E12:E14)"
    ws[f"F{gt_row}"] = "=SUM(F12:F14)"
    for col in ("D", "E", "F"):
        ws[f"{col}{gt_row}"].fill = TOTAL_FILL
        ws[f"{col}{gt_row}"].font = TOTAL_FONT
    _format_imps_cell(ws[f"D{gt_row}"])
    _format_imps_cell(ws[f"E{gt_row}"])
    _format_money_cell(ws[f"F{gt_row}"])

    ws[f"B{gt_row + 2}"] = (
        "Paste inventory export from TTD/DOOH platform into the DOOH Screenlist tab. "
        "Update rows 12–14 above with rolled-up venue counts and avails."
    )
    ws[f"B{gt_row + 2}"].font = NOTE_FONT
    ws.merge_cells(f"B{gt_row + 2}:F{gt_row + 2}")

    return ws


def build_dooh_screenlist(wb: Workbook) -> Worksheet:
    """DOOH Screenlist — pasted in by Ad Ops from the inventory platform."""
    ws = wb.create_sheet("DOOH Screenlist")
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables
    headers = [
        "id", "name", "asset_name", "publisher_code", "status", "integration_type",
        "publisher", "network", "media_type", "venue_type", "venue_name", "city",
        "postal_code", "county", "state", "dma", "country", "latitude", "longitude",
        "address", "four_week_impressions", "impression_multiplier", "bidstream_impressions",
        "screen_count", "min_duration", "max_duration", "supports_video", "supports_banner",
        "supports_audio", "slot_dimensions", "preferred_ad_format", "ad_formats",
        "restrictions", "languages", "asset_image_url", "programmatic_platform_key",
        "programmatic_platform_industry_key",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.value = h
        c.font = H_HEADER
        c.fill = H_HEADER_FILL
        c.alignment = CENTER
        c.border = HEADER_BORDER
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    ws["B3"] = (
        "Paste the inventory export from TTD / your DOOH platform below row 1. "
        "The Summary tab will roll up by Venue Type + City."
    )
    ws["B3"].font = NOTE_FONT

    return ws


def build_process_faqs(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet("Process FAQs")
    ws.sheet_view.showGridLines = False  # clean white margins outside the boxed tables
    widths = {"A": 1.5, "B": 2.0, "C": 36, "D": 60, "E": 50}
    _apply_col_widths(ws, widths)
    # No customer billing block here — this is a reference/FAQ tab, not a
    # billable line-item sheet, so a "Customer Billing Information" block in
    # column F has nothing to bill against and doesn't apply.
    _write_meta_block(ws, "Process FAQs", include_billing=False)

    _set_header(ws, 6, [
        ("C", "SERVICES"),
        ("D", "Frequently Asked Question"),
        ("E", "Resource Link"),
    ])

    faqs = [
        ("All Products",
         "Where can I see the specs and best practices needed for my digital product?",
         "Entravision's Digital Specs and Best Practices"),
        ("All Products",
         "Why should I install Entravision's GTM (Google Tag Manager) on my website?",
         "Entravision's GTM Process"),
        ("All Products",
         "How can I install the provided GTM in my CMS?",
         "Entravision's GTM Process for Most Common CMS Website Providers"),
        ("Facebook & Instagram Ads on EVC Radio | Noticias Ya Pages",
         "How do I accept an Entravision Meta Business Account request to initiate a brand partnership or post a co-branded ad?",
         "How to Meta partnership for co-branded ads guide"),
        ("TikTok Ads (Client's Handle) or TikTok Ads on EVC Radio | Noticias Ya Pages",
         "How can I provide access to my TikTok Page to Run Ads?",
         "How to TikTok Access Guide"),
        ("TikTok Ads",
         "How can I authorize Entravision's TikTok Business Center account via QR code?",
         "How to TikTok Business Center Entravision Account (QR Code Method)"),
    ]
    row = 8
    for svc, q, link in faqs:
        ws[f"C{row}"] = svc
        ws[f"C{row}"].font = BODY_BOLD
        ws[f"C{row}"].alignment = LEFT
        ws[f"C{row}"].border = THIN_BORDER
        ws[f"D{row}"] = q
        ws[f"D{row}"].font = BODY_FONT
        ws[f"D{row}"].alignment = LEFT
        ws[f"D{row}"].border = THIN_BORDER
        ws[f"E{row}"] = link
        ws[f"E{row}"].font = Font(name="Arial", size=10, color="FF0066CC", underline="single")
        ws[f"E{row}"].alignment = LEFT
        ws[f"E{row}"].border = THIN_BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    ws[f"C{row+1}"] = "CATEGORY AND PRODUCT RESTRICTIONS"
    ws[f"C{row+1}"].font = BODY_BOLD
    ws[f"C{row+1}"].fill = SUBTOTAL_FILL
    ws.merge_cells(f"C{row+1}:E{row+1}")

    cats = [
        ("Healthcare Advertisers — Google Search and YouTube Ads",
         "Filling out the Google healthcare advertising form is mandatory.",
         "Apply for healthcare-related advertising"),
        ("Consumer Finance Advertisers — Google Search and YouTube Ads",
         "Related to: 36% APR, state licenses for Financial Verification, APR & Repayment disclosures above the fold, and \"Instant\" claims.",
         "Consumer Finance Google Advertising Guidelines"),
        ("Government and Advocacy — Google Search and YouTube Ads",
         "Related to: Official Contracts, State License, SOS Filing, direct URL to government directories.",
         "Government Google Advertising Guidelines"),
        ("Alcohol — All products",
         "Related to: Alcohol or alcohol consumption.",
         "FTC & TTB Alcohol Advertising Guidelines"),
        ("Loans — All products",
         "Related to: Truth-in-lending compliance and disclosure requirements.",
         "FTC Advertising Compliance Guide - Loan Campaigns"),
    ]
    row += 2
    for svc, desc, link in cats:
        ws[f"C{row}"] = svc
        ws[f"C{row}"].font = BODY_BOLD
        ws[f"C{row}"].alignment = LEFT
        ws[f"D{row}"] = desc
        ws[f"D{row}"].alignment = LEFT
        ws[f"E{row}"] = link
        ws[f"E{row}"].font = Font(name="Arial", size=10, color="FF0066CC", underline="single")
        ws.row_dimensions[row].height = 30
        row += 1

    return ws


def build_workbook(output_path: Path, families_to_include=None) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    products = CATALOG
    if families_to_include:
        products = [p for p in CATALOG if p.family in families_to_include]

    build_proposal_a(wb, products, with_sections=False)
    build_proposal_a(wb, products, with_sections=True)
    build_proposal_a_gross(wb, products)
    build_avails_only(wb, products)
    build_dooh_summary(wb)
    build_dooh_screenlist(wb)
    build_process_faqs(wb)

    wb.save(output_path)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="Entravision_Digital_Media_Proposal_Template.xlsx")
    args = ap.parse_args()
    out = Path(args.out)
    build_workbook(out)
    print(f"Wrote: {out.resolve()}")

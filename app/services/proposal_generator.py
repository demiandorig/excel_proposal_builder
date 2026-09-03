"""
Proposal generator — produces a populated Excel proposal from a ProposalRequest
plus planner-curated line items.

Reuses the Stage 1 template builders, then overrides the meta block with
real client/campaign info and overrides each product row's budget/months/notes
with the planner's values.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.catalog import CATALOG, Product, by_name, families, by_family
from app.services.notion_parser import ProposalRequest, classify_output_tabs
from app import excel_template as et


# ---------------------------------------------------------------------------
# Planner-curated line item
# ---------------------------------------------------------------------------

@dataclass
class LineItem:
    """
    One row in the proposal — the planner picks a catalog product and assigns it
    a budget. months/notes override defaults; rate_override is for KLYY-style
    custom CPMs flagged in the business logic spec.
    """
    product_name: str               # canonical catalog name
    monthly_budget: float           # net dollars per month
    months: int = 3                 # flight length in months
    rate_override: Optional[float] = None     # optional CPM/CPP override
    notes_override: Optional[str] = None      # appended to the catalog note
    target_override: Optional[str] = None     # column D — defaults to campaign target
    # Secondary audience for added scale/avails (e.g. a broader look-alike
    # segment layered on top of a narrow primary intent audience). When
    # set, column D becomes a two-line "Primary: .. / Secondary: .." cell
    # instead of the single target line.
    target_secondary: Optional[str] = None
    # Stable per-line identity from the frontend. Two lines can share the same
    # product_name (e.g. same product, different targeting) — avails_data is
    # keyed by this id, NOT product_name, so those lines don't collide.
    id: Optional[str] = None
    # Per-line override of the catalog's estimated CPM (Fixed/impressions-
    # estimate products only — Meta, YouTube, TikTok, LinkedIn, Spotify,
    # Branded Content, ...). Distinct from rate_override, which is a real
    # CPM/CPP billing rate: this only feeds the "Est. $" impressions-vs-
    # spend calculation, never an actual charge. None = use the catalog's
    # own estimated_cpm_for_imps.
    estimated_cpm_override: Optional[float] = None
    # Added Value: a $0 (or below-minimum) budget is valid and expected for
    # this line, not a planner oversight — exempts it from the below-
    # minimum validation highlight in the app, and sorts it to the bottom
    # of the export's line-items table (above the totals row, below every
    # paid line) rather than interleaved among real budget lines.
    is_added_value: bool = False

    def total_budget(self) -> float:
        return self.monthly_budget * self.months


@dataclass
class AddonItem:
    """
    One planner-picked extra from Step 04's Add-Ons module — a fixed-price,
    one-time line (landing page, call tracking, brand lift study, ...),
    distinct from LineItem: no months/target/rate, just a name and an
    editable flat amount. Rolled into the proposal's ADD-ONS / ONE-TIME
    FEES export block instead of the main line-items table.
    """
    product_name: str               # canonical catalog name (a catalog.Product with is_addon=True)
    amount: float                   # planner-edited flat price; defaults to the catalog minimum_spend client-side
    notes_override: Optional[str] = None


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_proposal(
    request: ProposalRequest,
    line_items: list[LineItem],
    output_path: Path,
    *,
    force_tabs: Optional[dict] = None,
    enrichment=None,        # ProposalEnrichment | None
    proposal_title: Optional[str] = None,
    avails_data: Optional[dict] = None,   # product_name -> {max_imps, max_spend, est_uniques}
    tiers: Optional[list[dict]] = None,   # [{"label": "A", "line_items": [...], "avails_data": {...}}, ...]
    addons: Optional[list[AddonItem]] = None,   # Step 04's Add-Ons module picks — proposal-wide, not per-tier
) -> dict:
    """
    Generate an Excel proposal for the given request + line items.

    Args:
        request: parsed ProposalRequest (from notion_parser.parse_notion)
        line_items: planner-confirmed products with budgets — the single-tier
                    (legacy) call shape; ignored when `tiers` is given
        output_path: where to write the .xlsx
        force_tabs: optionally override the auto-classified tab set
                    (lets the planner force Gross, force wsections, etc.)
        tiers: tiered-budget options (up to 4). Each requested tab type
               (Net/wsections/Gross/Avails-Only) is built once PER TIER,
               lettered "Proposal A", "Proposal B", ... — the same lettering
               convention the single-tier sheet name already used, extended
               to one tab per option. DOOH and Process FAQs stay single,
               shared tabs regardless of tier count (they're inventory/
               reference content, not a budget scenario). When omitted or a
               single tier, output is byte-for-byte the same as before this
               parameter existed.
        addons: Step 04's Add-Ons picks (fixed-price extras — landing pages,
                call tracking, etc.) — the SAME list appears on every tier's
                sheet, matching how the old hardcoded add-ons list already
                behaved before it became planner-driven. None falls back to
                that legacy hardcoded list; [] means "planner picked none."

    Returns:
        dict with summary: {tabs_built: [...], total_net: float,
        total_gross: float, warnings: [...], tiers: [...] | None}. `tiers`
        (present only when there's more than one) carries each option's own
        {label, total_net, total_gross}; the top-level total_net/total_gross
        always describe the FIRST option, matching what a single-tier
        proposal's totals have always meant.
    """
    if not tiers:
        tiers = [{"label": "A", "line_items": line_items, "avails_data": avails_data}]
    multi_tier = len(tiers) > 1

    # Convert AddonItems into the plain dicts excel_template expects — kept
    # dict-based at that boundary (like avails_data already is) so
    # excel_template.py doesn't need to import this module's dataclasses.
    addons_dicts: Optional[list[dict]] = None
    if addons is not None:
        addons_dicts = []
        for a in addons:
            p = by_name(a.product_name)
            desc = (p.proposal_description if p else "") or ""
            if a.notes_override:
                desc = f"{desc}\n— {a.notes_override}" if desc else a.notes_override
            addons_dicts.append({
                "name": a.product_name,
                "description": desc,
                "amount": a.amount,
                "type": "Added Value" if not a.amount else "Fixed",
            })

    # 1. Resolve which tabs to build — a single decision shared by every tier
    #    (they're all the same request/request_type; only the product mix and
    #    budgets differ), based on the UNION of products across all tiers so
    #    a DOOH product tucked into option C still gets its DOOH tabs.
    all_product_names = [li.product_name for t in tiers for li in t["line_items"]]
    tabs = classify_output_tabs(
        request.request_type,
        all_product_names,
        has_agency_fee=request.agency_fee is not None and request.agency_fee > 0,
    )
    if force_tabs:
        tabs.update(force_tabs)

    # Build a blurb lookup from enrichment (product_name → blurb text) — a
    # product's blurb doesn't depend on which tier it appears in, so this is
    # shared across every tier's sheets.
    blurbs: dict[str, str] = {}
    if enrichment and enrichment.product_blurbs:
        for pb in enrichment.product_blurbs:
            if pb.product_name and pb.blurb:
                blurbs[pb.product_name] = pb.blurb

    wb = Workbook()
    wb.remove(wb.active)

    tabs_built = []
    warnings: list[str] = list(request.warnings)
    start_date = request.start_date or ""
    end_date = request.end_date or ""
    total_months = request.total_months or 3
    campaign_name = getattr(enrichment, "campaign_name", "") if enrichment else ""

    tier_summaries = []

    for tier in tiers:
        label = (tier.get("label") or "A").strip() or "A"
        tier_avails = tier.get("avails_data")
        tier_title_suffix = f" — Option {label}" if multi_tier else ""

        # Materialize Product objects in the order the planner provided them.
        # Filter products and line_items together so a skipped/unmatched
        # product can't shift the pairing between the two lists (they must
        # stay index-aligned for every zip()/index-based lookup downstream).
        products: list[Product] = []
        matched_line_items: list[LineItem] = []
        for li in tier["line_items"]:
            p = by_name(li.product_name)
            if p is None:
                warn_prefix = f"Option {label}: " if multi_tier else ""
                warnings.append(f"{warn_prefix}Product '{li.product_name}' not found in catalog — skipping.")
                continue
            products.append(p)
            matched_line_items.append(li)
        tier_line_items = matched_line_items

        # Added Value lines sort to the bottom of the EXPORT (above the
        # totals row), regardless of how the planner ordered them while
        # curating — that ordering is fully planner-controlled (Step 04's
        # drag-to-reorder) and shouldn't be second-guessed for the
        # interactive table; a finished proposal document just reads
        # better with $0/bundled extras grouped at the end rather than
        # interleaved among real paid lines. Stable sort — only moves
        # Added Value items past non-Added-Value ones, doesn't otherwise
        # reorder within either group.
        if any(li.is_added_value for li in tier_line_items):
            reordered = sorted(zip(products, tier_line_items), key=lambda pair: pair[1].is_added_value)
            products = [p for p, _ in reordered]
            tier_line_items = [li for _, li in reordered]

        if tabs.get("net"):
            ws = et.build_proposal_a(wb, products, with_sections=False,
                                     start_date=start_date, end_date=end_date, total_months=total_months,
                                     sheet_name=f"Proposal {label}", addons=addons_dicts)
            _populate_meta(ws, request, gross=False, proposal_title=proposal_title,
                           campaign_name=campaign_name, title_suffix=tier_title_suffix)
            _populate_line_items(ws, products, tier_line_items, gross=False, blurbs=blurbs,
                                 avails_data=tier_avails, request=request)
            tabs_built.append(f"Proposal {label}")

        if tabs.get("wsections"):
            ws = et.build_proposal_a(wb, products, with_sections=True,
                                     start_date=start_date, end_date=end_date, total_months=total_months,
                                     sheet_name=f"Proposal {label} (wsections)", addons=addons_dicts)
            _populate_meta(ws, request, gross=False, proposal_title=proposal_title,
                           campaign_name=campaign_name, title_suffix=tier_title_suffix)
            _populate_line_items(ws, products, tier_line_items, gross=False, with_sections=True, blurbs=blurbs,
                                 avails_data=tier_avails, request=request)
            tabs_built.append(f"Proposal {label} (wsections)")

        if tabs.get("gross"):
            ws = et.build_proposal_a_gross(wb, products,
                                           start_date=start_date, end_date=end_date, total_months=total_months,
                                           sheet_name=f"Proposal {label} (Gross)", addons=addons_dicts)
            _populate_meta(ws, request, gross=True, proposal_title=proposal_title,
                           campaign_name=campaign_name, title_suffix=tier_title_suffix)
            _populate_line_items(ws, products, tier_line_items, gross=True, blurbs=blurbs,
                                 avails_data=tier_avails, request=request)
            # Set agency fee in I14 (Gross sheet's variable input cell)
            if request.agency_fee is not None:
                ws["I14"] = request.agency_fee
                ws["I14"].number_format = "0.00%"
                ws["I14"].font = Font(name="Arial", size=10, bold=True, color="FF0000FF")
                ws["I14"].fill = PatternFill("solid", start_color="FFFFF2CC")
            tabs_built.append(f"Proposal {label} (Gross)")

        if tabs.get("avails_only"):
            avails_sheet_name = f"Avails-Only {label}" if multi_tier else "Avails-Only"
            ws = et.build_avails_only(wb, products, line_items=tier_line_items, request=request,
                                      start_date=start_date, end_date=end_date, avails_data=tier_avails,
                                      campaign_name=campaign_name, sheet_name=avails_sheet_name)
            _populate_meta(ws, request, gross=False, proposal_title=proposal_title,
                           title_suffix=f" (Avails-Only){tier_title_suffix}",
                           include_billing=False, include_campaign_meta=False)
            tabs_built.append(avails_sheet_name)

        tier_total_net = sum(li.total_budget() for li in tier_line_items)
        fee = request.agency_fee or 0.0
        tier_summaries.append({
            "label": label,
            "total_net": tier_total_net,
            "total_gross": tier_total_net / (1 - fee) if fee else tier_total_net,
        })

    # DOOH / Process FAQs — shared, single instance regardless of tier count
    if tabs.get("dooh_summary"):
        ws = et.build_dooh_summary(wb)
        # DOOH summary has its own structure; just stamp the client name in title
        ws["A1"] = f"DOOH Summary — {request.client_name or 'TBD'}"
        tabs_built.append("DOOH Summary")

    if tabs.get("dooh_screenlist"):
        et.build_dooh_screenlist(wb)
        tabs_built.append("DOOH Screenlist")

    # Always include Process FAQs as a reference tab
    et.build_process_faqs(wb)
    tabs_built.append("Process FAQs")

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    base = tier_summaries[0] if tier_summaries else {"total_net": 0.0, "total_gross": 0.0}
    return {
        "tabs_built": tabs_built,
        "total_net": base["total_net"],
        "total_gross": base["total_gross"],
        "tiers": tier_summaries if multi_tier else None,
        "output_path": str(output_path),
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Meta block & line item population (overwrites template defaults)
# ---------------------------------------------------------------------------

def _populate_meta(
    ws,
    request: ProposalRequest,
    *,
    gross: bool,
    title_suffix: str = "",
    proposal_title: Optional[str] = None,
    include_billing: bool = True,
    include_campaign_meta: bool = True,
    campaign_name: str = "",
) -> None:
    """
    Overwrite the meta block cells (rows 4-15) with real client info.
    Layout matches what _write_meta_block in excel_template.py established.

    include_campaign_meta: the C11..C14 "Media Proposal / Order Description /
    Geo" lines only apply to the standard Proposal A layout, where product
    rows start at row 19. The Avails-Only sheet uses a different layout
    (product rows start at row 11) and writes its own compact meta at
    C5/C6/C7 — writing here too would overwrite its first few product rows,
    so callers for that sheet must pass False.

    campaign_name: the AI-inferred short campaign name (e.g. "Summer
    Specials June 2026") written into the "Order Description:" line. Per-line
    targeting now lives solely in column D of each product row — there's no
    separate campaign-level "Target:" meta line anymore.
    """
    # Title cell (E2 — merged E2:L2)
    if proposal_title:
        title = proposal_title + (title_suffix if title_suffix else "")
    else:
        title = f"Digital Media Proposal — {request.client_name or 'TBD'}{title_suffix}"
    ws["E2"] = title

    # Customer billing block (F4..F8) — skipped on Avails-Only (not relevant there)
    if include_billing:
        ws["F5"] = f"Attn: {request.requested_by or ''}"
        ws["F6"] = f"Billing Name: {request.agency_name or request.client_name or ''}"
        ws["F7"] = f"Contact e-mail: {request.salesperson_email or ''}"
        # F8 (Address) intentionally left blank for planner to fill

    if not include_campaign_meta:
        return

    # Campaign meta block (C11..C15) — Proposal A layout only
    media_proposal_line = f"Media Proposal: {request.client_name or 'TBD'}"
    if request.start_date and request.end_date:
        media_proposal_line += f" — {request.start_date} to {request.end_date}"
    elif request.renewal_campaign_dates:
        media_proposal_line += f" — {request.renewal_campaign_dates}"
    ws["C11"] = media_proposal_line

    ws["C12"] = f"Order Description: {campaign_name}" if campaign_name else "Order Description: "

    if request.geo:
        ws["C13"] = f"Geo: {request.geo}"

    # "Minimum 3 month Commitment" only applies (and should only be shown)
    # when the actual flight is 3+ months — _write_meta_block's default
    # text stated it unconditionally regardless of the real flight length.
    if (request.total_months or 0) >= 3:
        ws["C14"] = "All rates are NET. Minimum 3 month Commitment."
    else:
        ws["C14"] = "All rates are NET."
    ws["C14"].font = et.BODY_BOLD


def _populate_line_items(
    ws,
    products: list[Product],
    line_items: list[LineItem],
    *,
    gross: bool,
    with_sections: bool = False,
    blurbs: Optional[dict] = None,
    avails_data: Optional[dict] = None,
    request: Optional[ProposalRequest] = None,
) -> None:
    """
    Walk the product rows already written by build_proposal_a / _gross and
    overwrite L (NET BUDGET), K (NET RATE if rate_override given), and the
    notes column with planner's values.

    The template builders write product rows starting at PRODUCT_START_ROW.
    For with_sections=True they intersperse section banners — we walk the
    same way to land on the right rows.
    """
    # Row layout starts at row 19 (after meta + header rows) — must match
    # PRODUCT_START_ROW used by excel_template.build_proposal_a.
    # In with_sections mode, the banner takes row N and the product takes N+1.
    start_row = 19

    row = start_row
    first_data_row = start_row
    last_family = None
    sov_col = "S" if gross else "Q"
    for product, li in zip(products, line_items):
        if with_sections and product.family != last_family:
            # Section banner just landed on `row`; product row is next.
            row += 1
            last_family = product.family

        # NET BUDGET (L) — the planner's MONTHLY budget, not the flight total.
        # Every other formula on this sheet assumes that: "TOTAL DIGITAL
        # MONTHLY" is SUM(L), and the grand total then multiplies that by
        # the months cell (I10) to get the flight total. Writing the flight
        # total here instead (monthly × months) double-counts months in
        # the grand total, and mislabels the monthly total 3x too high for
        # a 3-month flight — the exact bug the planner reported.
        ws[f"L{row}"] = li.monthly_budget
        et._format_money_cell(ws[f"L{row}"], blue_input=True)

        # NET RATE (K) — override only if planner specified a custom rate
        if li.rate_override is not None:
            ws[f"K{row}"] = li.rate_override
            et._format_money_cell(ws[f"K{row}"], blue_input=True)

        # TARGET (D) — per-line override, else Demo | Behavioral | Contextual
        # fallback (instead of leaving the template's default "TBD"); a
        # secondary audience turns this into a two-line Primary/Secondary
        # rich-text cell with bolded labels.
        ws[f"D{row}"] = et.build_target_cell_value(li.target_override, li.target_secondary, request)

        # AI BLURB (E) — overrides catalog proposal_description when present
        if blurbs and product.name in blurbs:
            ws[f"E{row}"] = blurbs[product.name]
            # Re-estimate row height for the (usually longer) AI-written text
            ws.row_dimensions[row].height = et._estimate_row_height(blurbs[product.name], col_width=50)

        # Notes column — T for net, W for gross
        notes_col = "W" if gross else "T"
        note_parts = [product.notes or ""]
        if li.is_added_value:
            # A $0 (or below-minimum) budget on this line is deliberate, not
            # an oversight — flag it inline so a reader doesn't mistake it
            # for a data error.
            note_parts.append("Added Value — no media cost.")
        if li.notes_override:
            note_parts.append(li.notes_override)
        combined = "\n— ".join(p for p in note_parts if p)
        if combined:
            ws[f"{notes_col}{row}"] = combined

        # Avails (planner-entered from Step 06 of the app) — N/O/P net, P/Q/R gross,
        # plus the SOV% column right after (Q net, S gross). Always written,
        # even with an empty avails dict, so a line with nothing entered gets
        # write_avails_cells's grey "not entered" flag instead of the row
        # silently staying plain-blank with no visual signal either way.
        # Keyed by the line item's own id (falls back to product name for any
        # older/manual request that didn't send one) — id-keying is what lets
        # two lines with the same product carry independent avails.
        avails_by = avails_data or {}
        avail = avails_by.get(li.id) if li.id else None
        if avail is None:
            avail = avails_by.get(product.name)
        sov_pct = et.compute_sov_pct(product, li.monthly_budget, avail or {}, cpm_override=li.estimated_cpm_override)
        et.write_avails_cells(ws, row, avail or {}, product, gross=gross, sov_pct=sov_pct, sov_col=sov_col,
                              budget_col="L", cpm_override=li.estimated_cpm_override)

        row += 1

    if sov_col:
        et._apply_sov_conditional_formatting(ws, sov_col, first_data_row, row - 1)
